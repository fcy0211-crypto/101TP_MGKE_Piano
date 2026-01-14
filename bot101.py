import asyncio
import sqlite3
from datetime import datetime, timedelta

from aiogram import Bot, Dispatcher
from aiogram.types import (
    Message, CallbackQuery,
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
    FSInputFile
)
from aiogram.filters import Command

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ================= НАСТРОЙКИ =================
BOT_TOKEN = "8299815983:AAEm94SDXpUD1C8O6QtQ2MhsiydG3KPJEak"
ADMIN_USERNAME = "Glabak0200"

DB_NAME = "attendance.db"
EXCEL_NAME = "rapport.xlsx"

STUDENTS = [
    "Бабук Владислав","Гарцуев Ростислав","Глинская Милена",
    "Демьянко Надежда","Касьянюк Глеб","Мигутский Тимур",
    "Михальчик Илья","Полторако Артём","Русецкая Кристина",
    "Серяков Игорь","Шаболтас Матвей"
]

REASONS = [
    "по заявлению","по болезни","по неуважительной причине"
]

HOURS = [1, 2, 3, 4, 5, 6]

bot = Bot(BOT_TOKEN)
dp = Dispatcher()
ADMIN_CHAT_ID = None

# ================= БАЗА =================
def db():
    return sqlite3.connect(DB_NAME)

def init_db():
    with db() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            student TEXT,
            reason TEXT,
            hours INTEGER,
            author TEXT,
            deleted_at TEXT
        )
        """)
        con.commit()

def today():
    return datetime.now().strftime("%Y-%m-%d")

def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ================= EXCEL =================
def export_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Рапортичка"

    header_fill = PatternFill("solid", fgColor="DDDDDD")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    headers = ["Дата", "ФИО", "Статус", "Причина", "Часы", "Кто отметил"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = header_fill

    with db() as con:
        dates = con.execute("""
        SELECT DISTINCT date FROM attendance
        WHERE deleted_at IS NULL
        ORDER BY date
        """).fetchall()

    for (date,) in dates:
        with db() as con:
            rows = con.execute("""
            SELECT student, reason, hours, author
            FROM attendance
            WHERE date = ? AND deleted_at IS NULL
            """, (date,)).fetchall()

        absent = {r[0]: (r[1], r[2], r[3]) for r in rows}

        for student in sorted(STUDENTS):
            if student in absent:
                reason, hours, author = absent[student]
                ws.append([date, student, "отсутствовал", reason, hours, author])
                for cell in ws[ws.max_row]:
                    cell.fill = red_fill
            else:
                ws.append([date, student, "присутствовал", "", "", ""])
                for cell in ws[ws.max_row]:
                    cell.fill = green_fill

    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    wb.save(EXCEL_NAME)

# ================= КЛАВИАТУРА =================
def menu():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="➕ Отметить")],
            [KeyboardButton(text="📅 Отметить за прошлые даты")],
            [KeyboardButton(text="✏ Редактировать")],
            [KeyboardButton(text="📤 Выгрузить")],
            [KeyboardButton(text="📨 Админу")],
            [KeyboardButton(text="🗑 Очистить")],
            [KeyboardButton(text="♻ Восстановить")]
        ], resize_keyboard=True
    )

# ================= START =================
@dp.message(Command("start"))
async def start(msg: Message):
    global ADMIN_CHAT_ID
    if msg.from_user.username == ADMIN_USERNAME:
        ADMIN_CHAT_ID = msg.chat.id
        await msg.answer("✅ Ты администратор")
    await msg.answer("📘 Рапортичка 101 тп", reply_markup=menu())

# ================= ВЫБОР СТУДЕНТА =================
async def choose_student_for_date(msg, date_str):
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for i, s in enumerate(STUDENTS):
        kb.inline_keyboard.append([InlineKeyboardButton(text=s, callback_data=f"s_{date_str}|{i}")])
    await msg.answer(f"Дата: {date_str}", reply_markup=kb)

# ================= ОТМЕТКА НА СЕГОДНЯ =================
@dp.message(lambda m: m.text == "➕ Отметить")
async def mark_today(msg: Message):
    await choose_student_for_date(msg, today())

# ================= ОТМЕТКА ЗА ПРОШЛЫЕ ДАТЫ =================
@dp.message(lambda m: m.text == "📅 Отметить за прошлые даты")
async def mark_past(msg: Message):
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for i in range(1, 15):
        date_obj = datetime.now() - timedelta(days=i)
        date_str = date_obj.strftime("%Y-%m-%d")
        kb.inline_keyboard.append([InlineKeyboardButton(text=date_str, callback_data=f"d_{date_str}")])
    await msg.answer("Выберите дату:", reply_markup=kb)

# ================= ВЫБОР ДАТЫ =================
@dp.callback_query(lambda c: c.data.startswith("d_"))
async def select_date(call: CallbackQuery):
    date_str = call.data[2:]
    await choose_student_for_date(call.message, date_str)

# ================= ВЫБОР СТУДЕНТА =================
@dp.callback_query(lambda c: c.data.startswith("s_"))
async def select_student(call: CallbackQuery):
    _, rest = call.data.split("_")
    date_str, student_idx = rest.split("|")
    student_idx = int(student_idx)
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for i, r in enumerate(REASONS):
        kb.inline_keyboard.append([InlineKeyboardButton(text=r, callback_data=f"r_{date_str}|{student_idx}|{i}")])
    await call.message.answer(f"{STUDENTS[student_idx]}\nВыберите причину отсутствия:", reply_markup=kb)

# ================= ВЫБОР ПРИЧИНЫ =================
@dp.callback_query(lambda c: c.data.startswith("r_"))
async def select_reason(call: CallbackQuery):
    _, rest = call.data.split("_")
    date_str, student_idx, reason_idx = rest.split("|")
    student_idx = int(student_idx)
    reason_idx = int(reason_idx)
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for h in HOURS:
        kb.inline_keyboard.append([InlineKeyboardButton(text=f"{h} ч", callback_data=f"h_{date_str}|{student_idx}|{reason_idx}|{h}")])
    await call.message.answer("Сколько часов отсутствовал?", reply_markup=kb)

# ================= ВЫБОР ЧАСОВ =================
@dp.callback_query(lambda c: c.data.startswith("h_"))
async def select_hours(call: CallbackQuery):
    _, rest = call.data.split("_")
    date_str, student_idx, reason_idx, hours = rest.split("|")
    student_idx = int(student_idx)
    reason_idx = int(reason_idx)
    hours = int(hours)
    with db() as con:
        con.execute("""
        INSERT INTO attendance (date, student, reason, hours, author, deleted_at)
        VALUES (?, ?, ?, ?, ?, NULL)
        """, (
            date_str,
            STUDENTS[student_idx],
            REASONS[reason_idx],
            hours,
            call.from_user.username or call.from_user.full_name
        ))
        con.commit()
    await call.message.answer(f"✅ Отмечено: {date_str} | {STUDENTS[student_idx]} | {REASONS[reason_idx]} | {hours} ч")

# ================= РЕДАКТИРОВАНИЕ =================
@dp.message(lambda m: m.text == "✏ Редактировать")
async def edit(msg: Message):
    with db() as con:
        rows = con.execute("""
        SELECT id, date, student, reason, hours
        FROM attendance
        WHERE deleted_at IS NULL
        """).fetchall()
    if not rows:
        await msg.answer("Нет записей")
        return
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for r in rows:
        kb.inline_keyboard.append([InlineKeyboardButton(text=f"{r[1]} | {r[2]} ({r[4]} ч) | {r[3]}", callback_data=f"edit_{r[0]}")])
    await msg.answer("Выберите запись для редактирования:", reply_markup=kb)

@dp.callback_query(lambda c: c.data.startswith("edit_"))
async def edit_entry(call: CallbackQuery):
    rec_id = int(call.data[5:])
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for i, r in enumerate(REASONS):
        kb.inline_keyboard.append([InlineKeyboardButton(text=r, callback_data=f"editr_{rec_id}|{i}")])
    await call.message.answer("Выберите новую причину:", reply_markup=kb)

@dp.callback_query(lambda c: c.data.startswith("editr_"))
async def edit_reason_hours(call: CallbackQuery):
    rec_id, reason_idx = call.data[6:].split("|")
    reason_idx = int(reason_idx)
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for h in HOURS:
        kb.inline_keyboard.append([InlineKeyboardButton(text=f"{h} ч", callback_data=f"edith_{rec_id}|{reason_idx}|{h}")])
    await call.message.answer("Выберите новые часы:", reply_markup=kb)

@dp.callback_query(lambda c: c.data.startswith("edith_"))
async def update_reason_hours(call: CallbackQuery):
    rec_id, reason_idx, hours = call.data[6:].split("|")
    reason_idx = int(reason_idx)
    hours = int(hours)
    with db() as con:
        con.execute("UPDATE attendance SET reason=?, hours=? WHERE id=?", (REASONS[reason_idx], hours, int(rec_id)))
        con.commit()
    await call.message.answer("✏ Запись обновлена успешно!")

# ================= ВЫГРУЗКА =================
@dp.message(lambda m: m.text == "📤 Выгрузить")
async def export(msg: Message):
    export_excel()
    await msg.answer_document(FSInputFile(EXCEL_NAME))

@dp.message(lambda m: m.text == "📨 Админу")
async def send_admin(msg: Message):
    global ADMIN_CHAT_ID
    if not ADMIN_CHAT_ID:
        await msg.answer("Админ не активен")
        return
    export_excel()
    await bot.send_document(ADMIN_CHAT_ID, FSInputFile(EXCEL_NAME), caption="📊 Рапортичка")
    await msg.answer("✅ Отправлено")

@dp.message(lambda m: m.text == "🗑 Очистить")
async def clear(msg: Message):
    with db() as con:
        con.execute("UPDATE attendance SET deleted_at=? WHERE deleted_at IS NULL", (now(),))
        con.commit()
    await msg.answer("🗑 Очищено (восстановимо 30 дней)")

@dp.message(lambda m: m.text == "♻ Восстановить")
async def restore(msg: Message):
    limit = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")
    with db() as con:
        con.execute("UPDATE attendance SET deleted_at=NULL WHERE deleted_at >= ?", (limit,))
        con.commit()
    await msg.answer("♻ Восстановлено")

# ================= ЗАПУСК =================
async def main():
    init_db()
    print("Бот запущен")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
